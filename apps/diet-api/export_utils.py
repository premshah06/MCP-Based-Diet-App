#!/usr/bin/env python3
"""
Excel Export Utilities for Diet Coach System
Generates professional, formatted Excel reports for nutrition data
"""
import pandas as pd
from datetime import datetime
from typing import Dict, Any, List, Optional
import io
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
import base64

logger = logging.getLogger(__name__)

class NutritionExcelExporter:
    """Professional Excel report generator for nutrition data"""
    
    def __init__(self):
        # Define consistent styling
        self.header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        self.title_font = Font(name='Calibri', size=16, bold=True, color='2F75B5')
        self.data_font = Font(name='Calibri', size=11)
        self.small_font = Font(name='Calibri', size=9, color='666666')
        
        self.header_fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
        self.alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        self.success_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
        self.warning_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
        self.danger_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
    
    def generate_comprehensive_report(self, 
                                    user_profile: Dict[str, Any],
                                    nutrition_targets: Dict[str, Any],
                                    meal_plan: Dict[str, Any],
                                    explanation: str,
                                    validation_results: Optional[Dict] = None) -> bytes:
        """Generate comprehensive nutrition report in Excel format"""
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create worksheets
        self._create_summary_sheet(wb, user_profile, nutrition_targets, explanation)
        self._create_meal_plan_sheet(wb, meal_plan)
        self._create_nutrition_analysis_sheet(wb, nutrition_targets, meal_plan)
        self._create_food_database_sheet(wb, meal_plan)
        
        if validation_results:
            self._create_validation_sheet(wb, validation_results)
        
        self._create_guidelines_sheet(wb)
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
    
    def _create_summary_sheet(self, wb: Workbook, user_profile: Dict, nutrition_targets: Dict, explanation: str):
        """Create executive summary sheet"""
        ws = wb.create_sheet("Executive Summary", 0)
        
        # Title
        ws['A1'] = "PERSONALIZED NUTRITION REPORT"
        ws['A1'].font = Font(name='Calibri', size=20, bold=True, color='2F75B5')
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = self.center_alignment
        
        # Report metadata
        ws['A3'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws['A3'].font = self.small_font
        ws['A4'] = "Diet Coach AI - Professional Nutrition Analysis"
        ws['A4'].font = self.small_font
        
        # User profile section
        row = 6
        ws[f'A{row}'] = "CLIENT PROFILE"
        ws[f'A{row}'].font = self.title_font
        row += 2
        
        profile_data = [
            ("Age", f"{user_profile.get('age', 'N/A')} years"),
            ("Gender", user_profile.get('sex', 'N/A').title()),
            ("Height", f"{user_profile.get('height_cm', 'N/A')} cm"),
            ("Weight", f"{user_profile.get('weight_kg', 'N/A')} kg"),
            ("Activity Level", user_profile.get('activity_level', 'N/A').replace('_', ' ').title()),
            ("Goal", user_profile.get('goal', 'N/A').title()),
            ("BMI", f"{self._calculate_bmi(user_profile):.1f}" if self._calculate_bmi(user_profile) else "N/A")
        ]
        
        for label, value in profile_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
            ws[f'B{row}'].font = self.data_font
            row += 1
        
        # Nutrition targets section
        row += 2
        ws[f'A{row}'] = "DAILY NUTRITION TARGETS"
        ws[f'A{row}'].font = self.title_font
        row += 2
        
        # Create nutrition targets table
        headers = ['Nutrient', 'Target Amount', 'Unit', 'Calories from Nutrient', '% of Total Calories']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        row += 1
        total_calories = nutrition_targets.get('target_calories', 0)
        
        nutrition_data = [
            ('Calories', total_calories, 'kcal', total_calories, 100),
            ('Protein', nutrition_targets.get('macro_targets', {}).get('protein_g', 0), 'g', 
             nutrition_targets.get('macro_targets', {}).get('protein_g', 0) * 4,
             (nutrition_targets.get('macro_targets', {}).get('protein_g', 0) * 4 / total_calories * 100) if total_calories > 0 else 0),
            ('Fat', nutrition_targets.get('macro_targets', {}).get('fat_g', 0), 'g',
             nutrition_targets.get('macro_targets', {}).get('fat_g', 0) * 9,
             (nutrition_targets.get('macro_targets', {}).get('fat_g', 0) * 9 / total_calories * 100) if total_calories > 0 else 0),
            ('Carbohydrates', nutrition_targets.get('macro_targets', {}).get('carbs_g', 0), 'g',
             nutrition_targets.get('macro_targets', {}).get('carbs_g', 0) * 4,
             (nutrition_targets.get('macro_targets', {}).get('carbs_g', 0) * 4 / total_calories * 100) if total_calories > 0 else 0)
        ]
        
        for i, (nutrient, amount, unit, cal_from_nutrient, percentage) in enumerate(nutrition_data):
            for col, value in enumerate([nutrient, f"{amount:.1f}", unit, f"{cal_from_nutrient:.0f}", f"{percentage:.1f}%"], 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.data_font
                cell.border = self.thin_border
                cell.alignment = self.center_alignment if col > 1 else self.left_alignment
                if i % 2 == 1:
                    cell.fill = self.alt_fill
            row += 1
        
        # AI Dietitian Explanation
        row += 2
        ws[f'A{row}'] = "PROFESSIONAL DIETITIAN ANALYSIS"
        ws[f'A{row}'].font = self.title_font
        row += 2
        
        # Split explanation into paragraphs and format
        explanation_lines = explanation.split('\n')
        for line in explanation_lines:
            if line.strip():
                ws[f'A{row}'] = line.strip()
                ws[f'A{row}'].font = self.data_font
                ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                ws.merge_cells(f'A{row}:F{row}')
                ws.row_dimensions[row].height = 20
                row += 1
        
        # Auto-adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20
    
    def _create_meal_plan_sheet(self, wb: Workbook, meal_plan: Dict):
        """Create detailed meal plan sheet"""
        ws = wb.create_sheet("Meal Plan Details")
        
        # Title
        ws['A1'] = "DETAILED MEAL PLAN"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = self.center_alignment
        
        row = 3
        
        # Create meal plan table
        headers = ['Day', 'Meal', 'Food Item', 'Amount (g)', 'Calories', 'Protein (g)', 'Fat (g)', 'Carbs (g)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        row += 1
        
        for day in meal_plan.get('days', []):
            day_num = day.get('day', 1)
            first_row_of_day = row
            
            for meal in day.get('meals', []):
                meal_name = meal.get('name', 'Unknown Meal')
                first_row_of_meal = row
                
                for food in meal.get('foods', []):
                    data = [
                        day_num if row == first_row_of_day else '',
                        meal_name if row == first_row_of_meal else '',
                        food.get('name', 'Unknown Food'),
                        food.get('amount_g', 0),
                        food.get('calories', 0),
                        food.get('protein', 0),
                        food.get('fat', 0),
                        food.get('carbs', 0)
                    ]
                    
                    for col, value in enumerate(data, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.font = self.data_font
                        cell.border = self.thin_border
                        cell.alignment = self.center_alignment if col > 2 else self.left_alignment
                        if (day_num - 1) % 2 == 1:
                            cell.fill = self.alt_fill
                    row += 1
                
                # Add meal totals
                totals = meal.get('totals', {})
                total_data = ['', '', 'MEAL TOTAL:', '', 
                            totals.get('calories', 0),
                            totals.get('protein', 0),
                            totals.get('fat', 0),
                            totals.get('carbs', 0)]
                
                for col, value in enumerate(total_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = Font(name='Calibri', size=11, bold=True)
                    cell.border = self.thin_border
                    cell.alignment = self.center_alignment if col > 2 else self.left_alignment
                    if col >= 3:
                        cell.fill = self.success_fill
                row += 1
            
            # Add daily totals
            daily_totals = day.get('daily_totals', {})
            daily_data = ['', '', 'DAILY TOTAL:', '',
                         daily_totals.get('calories', 0),
                         daily_totals.get('protein', 0),
                         daily_totals.get('fat', 0),
                         daily_totals.get('carbs', 0)]
            
            for col, value in enumerate(daily_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
                cell.fill = self.header_fill
                cell.border = self.thin_border
                cell.alignment = self.center_alignment if col > 2 else self.left_alignment
            row += 2
        
        # Auto-adjust column widths
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col_letter].width = 15
        ws.column_dimensions['C'].width = 25  # Food names need more space
    
    def _create_nutrition_analysis_sheet(self, wb: Workbook, nutrition_targets: Dict, meal_plan: Dict):
        """Create nutrition analysis with charts"""
        ws = wb.create_sheet("Nutrition Analysis")
        
        # Title
        ws['A1'] = "NUTRITIONAL ANALYSIS & ADHERENCE"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = self.center_alignment
        
        # Calculate actual vs target
        plan_totals = meal_plan.get('plan_totals', {})
        targets = nutrition_targets.get('macro_targets', {})
        days = len(meal_plan.get('days', []))
        
        if days > 0:
            avg_daily_calories = plan_totals.get('calories', 0) / days
            avg_daily_protein = plan_totals.get('protein', 0) / days
            avg_daily_fat = plan_totals.get('fat', 0) / days
            avg_daily_carbs = plan_totals.get('carbs', 0) / days
        else:
            avg_daily_calories = avg_daily_protein = avg_daily_fat = avg_daily_carbs = 0
        
        # Create comparison table
        row = 3
        headers = ['Nutrient', 'Target', 'Actual (Avg Daily)', 'Difference', 'Adherence %', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        row += 1
        
        comparison_data = [
            ('Calories', nutrition_targets.get('target_calories', 0), avg_daily_calories, 'kcal'),
            ('Protein', targets.get('protein_g', 0), avg_daily_protein, 'g'),
            ('Fat', targets.get('fat_g', 0), avg_daily_fat, 'g'),
            ('Carbohydrates', targets.get('carbs_g', 0), avg_daily_carbs, 'g')
        ]
        
        for nutrient, target, actual, unit in comparison_data:
            difference = actual - target
            adherence = (actual / target * 100) if target > 0 else 0
            
            if 90 <= adherence <= 110:
                status = "Excellent"
                status_fill = self.success_fill
            elif 80 <= adherence <= 120:
                status = "Good"
                status_fill = self.warning_fill
            else:
                status = "Needs Adjustment"
                status_fill = self.danger_fill
            
            data = [
                nutrient,
                f"{target:.1f} {unit}",
                f"{actual:.1f} {unit}",
                f"{difference:+.1f} {unit}",
                f"{adherence:.1f}%",
                status
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.data_font
                cell.border = self.thin_border
                cell.alignment = self.center_alignment if col > 1 else self.left_alignment
                if col == 6:  # Status column
                    cell.fill = status_fill
            row += 1
        
        # Auto-adjust column widths
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col_letter].width = 18
    
    def _create_food_database_sheet(self, wb: Workbook, meal_plan: Dict):
        """Create food database reference sheet"""
        ws = wb.create_sheet("Food Database")
        
        # Title
        ws['A1'] = "FOOD DATABASE REFERENCE"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = self.center_alignment
        
        # Extract unique foods from meal plan
        unique_foods = {}
        for day in meal_plan.get('days', []):
            for meal in day.get('meals', []):
                for food in meal.get('foods', []):
                    food_name = food.get('name')
                    if food_name and food_name not in unique_foods:
                        unique_foods[food_name] = food
        
        # Create table
        row = 3
        headers = ['Food Name', 'Calories/100g', 'Protein/100g', 'Fat/100g', 'Carbs/100g', 'Category', 'Prep Time', 'Cost Level']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        row += 1
        
        for food_name, food_data in sorted(unique_foods.items()):
            # Calculate per 100g values
            amount_g = food_data.get('amount_g', 100)
            scale_factor = 100 / amount_g if amount_g > 0 else 1
            
            data = [
                food_name,
                f"{food_data.get('calories', 0) * scale_factor:.1f}",
                f"{food_data.get('protein', 0) * scale_factor:.1f}g",
                f"{food_data.get('fat', 0) * scale_factor:.1f}g",
                f"{food_data.get('carbs', 0) * scale_factor:.1f}g",
                self._categorize_food(food_name),
                "10-15 min",  # Default prep time
                "Medium"      # Default cost level
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.data_font
                cell.border = self.thin_border
                cell.alignment = self.center_alignment if col > 1 else self.left_alignment
                if row % 2 == 0:
                    cell.fill = self.alt_fill
            row += 1
        
        # Auto-adjust column widths
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col_letter].width = 15
        ws.column_dimensions['A'].width = 25
    
    def _create_validation_sheet(self, wb: Workbook, validation_results: Dict):
        """Create validation results sheet"""
        ws = wb.create_sheet("Quality Validation")
        
        # Title
        ws['A1'] = "QUALITY VALIDATION REPORT"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = self.center_alignment
        
        row = 3
        
        # Summary
        summary = validation_results.get('results', {}).get('summary', {})
        ws[f'A{row}'] = f"Overall Pass Rate: {summary.get('overall_pass_rate', 0)}%"
        ws[f'A{row}'].font = Font(name='Calibri', size=14, bold=True)
        row += 1
        
        ws[f'A{row}'] = f"Safety Approved: {'✓' if summary.get('safety_approved', False) else '✗'}"
        ws[f'A{row}'].font = self.data_font
        row += 1
        
        ws[f'A{row}'] = f"Cultural Sensitivity: {'✓' if summary.get('cultural_sensitivity_approved', False) else '✗'}"
        ws[f'A{row}'].font = self.data_font
        row += 2
        
        # Detailed results
        headers = ['Check Type', 'Result', 'Severity', 'Message', 'Recommendations']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        row += 1
        
        # Add validation details
        for category in ['nutrition', 'cultural', 'accessibility']:
            results = validation_results.get('results', {}).get(category, [])
            for result in results:
                data = [
                    result.get('check_name', '').replace('_', ' ').title(),
                    '✓ Pass' if result.get('passed', False) else '✗ Fail',
                    result.get('severity', {}).get('value', 'info').title(),
                    result.get('message', ''),
                    ', '.join(result.get('recommendations', [])[:2])  # First 2 recommendations
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = self.data_font
                    cell.border = self.thin_border
                    cell.alignment = self.left_alignment
                    if not result.get('passed', False):
                        cell.fill = self.warning_fill
                row += 1
        
        # Auto-adjust column widths
        for col_letter in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col_letter].width = 20
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 30
    
    def _create_guidelines_sheet(self, wb: Workbook):
        """Create nutritional guidelines reference sheet"""
        ws = wb.create_sheet("Guidelines & Tips")
        
        # Title
        ws['A1'] = "NUTRITIONAL GUIDELINES & TIPS"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = self.center_alignment
        
        row = 3
        
        guidelines = [
            ("General Guidelines", [
                "• Eat a variety of foods from all food groups",
                "• Aim for at least 5 servings of fruits and vegetables daily",
                "• Choose whole grains over refined grains",
                "• Include lean protein sources in each meal",
                "• Stay hydrated with 8-10 glasses of water daily",
                "• Limit processed foods and added sugars"
            ]),
            ("Meal Timing Tips", [
                "• Eat breakfast within 2 hours of waking",
                "• Space meals 3-4 hours apart",
                "• Have your largest meal when most active",
                "• Stop eating 2-3 hours before bedtime",
                "• Listen to hunger and fullness cues"
            ]),
            ("Portion Control", [
                "• Use smaller plates and bowls",
                "• Fill half your plate with vegetables",
                "• Protein should be palm-sized",
                "• Carbohydrates should be fist-sized",
                "• Fats should be thumb-sized",
                "• Eat slowly and mindfully"
            ]),
            ("Food Safety", [
                "• Wash hands before food preparation",
                "• Store perishables in refrigerator",
                "• Cook meats to safe temperatures",
                "• Avoid cross-contamination",
                "• Check expiration dates regularly"
            ])
        ]
        
        for title, tips in guidelines:
            ws[f'A{row}'] = title
            ws[f'A{row}'].font = Font(name='Calibri', size=14, bold=True, color='2F75B5')
            row += 1
            
            for tip in tips:
                ws[f'A{row}'] = tip
                ws[f'A{row}'].font = self.data_font
                ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                ws.merge_cells(f'A{row}:D{row}')
                row += 1
            row += 1
        
        # Auto-adjust column widths
        for col_letter in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col_letter].width = 25
    
    def _calculate_bmi(self, user_profile: Dict) -> Optional[float]:
        """Calculate BMI from user profile"""
        try:
            weight = user_profile.get('weight_kg')
            height = user_profile.get('height_cm')
            if weight and height:
                height_m = height / 100
                return weight / (height_m ** 2)
        except (TypeError, ZeroDivisionError):
            pass
        return None
    
    def _categorize_food(self, food_name: str) -> str:
        """Categorize food based on name"""
        food_name_lower = food_name.lower()
        
        if any(protein in food_name_lower for protein in ['chicken', 'fish', 'beef', 'turkey', 'tofu', 'egg']):
            return "Protein"
        elif any(carb in food_name_lower for carb in ['rice', 'pasta', 'bread', 'oat', 'quinoa']):
            return "Carbohydrate"
        elif any(veg in food_name_lower for veg in ['spinach', 'broccoli', 'carrot', 'pepper', 'kale']):
            return "Vegetable"
        elif any(fruit in food_name_lower for fruit in ['apple', 'banana', 'berry', 'orange']):
            return "Fruit"
        elif any(fat in food_name_lower for fat in ['oil', 'nut', 'avocado', 'seed']):
            return "Healthy Fat"
        else:
            return "Mixed/Other"

def create_excel_export(user_profile: Dict[str, Any],
                       nutrition_targets: Dict[str, Any],
                       meal_plan: Dict[str, Any],
                       explanation: str,
                       validation_results: Optional[Dict] = None) -> str:
    """Create Excel export and return base64 encoded string"""
    try:
        exporter = NutritionExcelExporter()
        excel_bytes = exporter.generate_comprehensive_report(
            user_profile, nutrition_targets, meal_plan, explanation, validation_results
        )
        return base64.b64encode(excel_bytes).decode('utf-8')
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        raise Exception(f"Failed to generate Excel report: {str(e)}")
